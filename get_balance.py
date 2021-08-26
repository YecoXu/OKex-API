import okex.Account_api as Account
import okex.Trade_api as Trade
import okex.Market_api as Market
from prettytable import PrettyTable
import okex.Funding_api as Funding
import datetime
import time
import json
import pandas as pd
import numpy as np
from openpyxl import Workbook
import openpyxl as op

usdt= 6.45

def init_basics():
    api_key = ""
    secret_key = ""
    passphrase = ""
    flag = '0'  # 实盘
    accountAPI = Account.AccountAPI(api_key, secret_key, passphrase, False, flag)
    tradeAPI = Trade.TradeAPI(api_key, secret_key, passphrase, False, flag)
    marketAPI = Market.MarketAPI(api_key, secret_key, passphrase, False, flag)
    fundingAPI = Funding.FundingAPI(api_key, secret_key, passphrase, False, flag)
    parameters = {"accountAPI": accountAPI,
                  "tradeAPI": tradeAPI,
                  "marketAPI":marketAPI,
                  "fundingAPI": fundingAPI
                  }
    return parameters

#转换unix时间戳日期毫秒格式为'%Y.%m.%d'
def timestamp_datetime(value):
    format = '%Y.%m.%d'
    temp = datetime.datetime.fromtimestamp(value / 1000).strftime(format)
    return temp


def add_coin_one(coinname):
    parameters = init_basics()
    tradeAPI = parameters["tradeAPI"]
    wb = op.load_workbook("coin_v5.xlsx")
    wb.create_sheet(coinname)

    result = tradeAPI.orders_history_archive(instType='SPOT', instId=coinname + '-USDT', state='filled')
    print(json.dumps(result['data'], sort_keys=True, indent=4, separators=(',', ': ')))
    if len(result['data']) == 0:
        return
    ws = wb[coinname]
    for re in result['data']:
        if re['side'] == "buy":
            ws.insert_rows(0)
            ws['A1'] = str(timestamp_datetime(int(re['fillTime'])))
            ws['B1'] = float(re['avgPx'])
            ws['C1'] = float(re['accFillSz']) * -1
            ws['D1'] = float(re['accFillSz']) * -1 * float(re['avgPx'])
        else:
            ws.insert_rows(0)
            ws['A1'] = str(timestamp_datetime(int(re['fillTime'])))
            ws['B1'] = float(re['avgPx'])
            ws['C1'] = float(re['accFillSz'])
            ws['D1'] = float(re['accFillSz']) * float(re['avgPx'])

    ws['H1'] = str(result['data'][0]['ordId'])
    print(coinname+"----------------------"+ws['H1'].value)
    wb.save("coin_v5.xlsx")

#把所有币的disEq放到对应sheet的E1位置  disEq  美金层面币种折算权益
def set_disEq_all():
    parameters = init_basics()
    accountAPI = parameters["accountAPI"]

    wb = op.load_workbook("coin_v5.xlsx")
    for coin_name in wb.sheetnames:
        result = accountAPI.get_account(coin_name)

        ws = wb[coin_name]
        if(len(result['data'][0]['details']))==0:
            ws['E1'] = 0
            continue
        details = result['data'][0]['details'][0]
        ws['E1'] = float(details['disEq'])
        print(coin_name + "------------------------------------------------------")
        print(details['disEq'])
    wb.save("coin_v5.xlsx")

#不导入v3数据 初始化
def init_account_order():
    parameters = init_basics()
    tradeAPI = parameters["tradeAPI"]
    wb = op.load_workbook("coin_v5.xlsx")
    for coin_name in wb.sheetnames:
      result = tradeAPI.orders_history_archive(instType='SPOT',instId=coin_name+'-USDT',state='filled')
      if len(result['data'])==0:
          continue
      ws=wb[coin_name]
      for re in result['data']:
          if re['side']=="buy":
              ws.insert_rows(0)
              ws['A1'] = str(timestamp_datetime(int(re['fillTime'])))
              ws['B1'] = float(re['avgPx'])
              ws['C1'] = float(re['accFillSz']) * -1
              ws['D1'] = float(re['accFillSz']) * -1*float(re['avgPx'])
          else:
              ws.insert_rows(0)
              ws['A1'] = str(timestamp_datetime(int(re['fillTime'])))
              ws['B1'] = float(re['avgPx'])
              ws['C1'] = float(re['accFillSz'])
              ws['D1'] = float(re['accFillSz'])  * float(re['avgPx'])

      ws['H1'] = str(result['data'][0]['ordId'])
    result = tradeAPI.orders_history_archive(instType='SPOT')

    wb['BTC']['I1']=result['data'][0]['ordId']#最新一条的订单信息

    wb.save("coin_v5.xlsx")
#导入v3数据 初始化
def init_account_order_v3():
    parameters = init_basics()
    tradeAPI = parameters["tradeAPI"]
    #t="302473049861935104"
    wb = op.load_workbook("coin_v5.xlsx")
    for c_name in wb.sheetnames:

        result = tradeAPI.orders_history_archive(instType='SPOT',instId=c_name+'-USDT',state='filled')
        if len(result['data']) == 0:
            continue
        ws = wb[c_name]
        n = ws.max_row
        result0_len = len(result['data'])
        for re in result['data']:
            if re['side'] == "buy":
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz']) * -1
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * -1 * float(re['avgPx'])

            else:
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz'])
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * float(re['avgPx'])
            n=n-1
        ws['H1'] = str(result['data'][0]['ordId'])

    result = tradeAPI.orders_history_archive(instType='SPOT')
    wb['BTC']['I1']=result['data'][0]['ordId']#最新一条的订单信息

    wb.save("coin_v5.xlsx")
#更新 单个币订单
def update_account_order_one(coin_name):
    parameters = init_basics()
    tradeAPI = parameters["tradeAPI"]
    wb = op.load_workbook("coin_v5.xlsx")
    ws=wb[coin_name]
    if ws['H1'].value is None:
        result = tradeAPI.orders_history_archive(instType='SPOT', instId=coin_name + '-USDT')
        ws['H1'] = str(result['data'][0]['ordId'])
        wb.save("coin_v5.xlsx")
        n = ws.max_row
        result0_len = len(result['data'])
        for re in result['data']:
            print(coin_name + "   Updating")
            wb = op.load_workbook("coin_v5.xlsx")
            ws = wb[coin_name]
            if re['side'] == "buy":
                print("*********************************")
                print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz']) * -1
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * -1 * float(re['avgPx'])

            else:
                print("*********************************")
                print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz'])
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * float(re['avgPx'])
            n = n - 1
            wb.save("coin_v5.xlsx")
    else:
        result = tradeAPI.orders_history_archive(instType='SPOT', instId=coin_name + '-USDT',
                                                 before=wb[coin_name]['H1'].value)
        ws['H1'] = str(result['data'][0]['ordId'])
        wb.save("coin_v5.xlsx")
        n = ws.max_row
        result0_len = len(result['data'])
        for re in result['data']:
            print(coin_name + "   Updating")
            if re['side'] == "buy":
                print("==================================")
                print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz']) * -1
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * -1 * float(re['avgPx'])
            else:
                print("==================================")
                print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                ws['B' + str(n + result0_len)] = float(re['avgPx'])
                ws['C' + str(n + result0_len)] = float(re['accFillSz'])
                ws['D' + str(n + result0_len)] = float(re['accFillSz']) * float(re['avgPx'])
            n = n - 1
            wb.save("coin_v5.xlsx")
    result = tradeAPI.orders_history_archive(instType='SPOT')
    wb['BTC']['I1'] = result['data'][0]['ordId']  # 最新一条的订单信息
    wb.save("coin_v5.xlsx")


#更新订单
def update_account_order_all():
    parameters = init_basics()
    tradeAPI = parameters["tradeAPI"]
    wb = op.load_workbook("coin_v5.xlsx")
    result = tradeAPI.orders_history_archive(instType='SPOT',before=wb['BTC']['I1'].value)
    coin_list=[]
    for re in result['data']:
        coin_name = re['instId'].split('-')[0]
        if coin_name in coin_list:
            continue
        coin_list.append(coin_name)
        if coin_name not in wb.sheetnames:#币 不在excel表中 添加币
            add_coin_one(coin_name)

        wb = op.load_workbook("coin_v5.xlsx")
        ws = wb[coin_name]

        if ws['H1'].value == None:
            result = tradeAPI.orders_history_archive(instType='SPOT', instId=coin_name + '-USDT')
            ws['H1'] = str(result['data'][0]['ordId'])
            wb.save("coin_v5.xlsx")
            n = ws.max_row
            result0_len = len(result['data'])
            for re in result['data']:
                print(coin_name + "   Updating")
                wb = op.load_workbook("coin_v5.xlsx")
                ws = wb[coin_name]
                if re['side'] == "buy":
                    print("*********************************")
                    print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                    ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                    ws['B' + str(n + result0_len)] = float(re['avgPx'])
                    ws['C' + str(n + result0_len)] = float(re['accFillSz']) * -1
                    ws['D' + str(n + result0_len)] = float(re['accFillSz']) * -1 * float(re['avgPx'])

                else:
                    print("*********************************")
                    print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                    ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                    ws['B' + str(n + result0_len)] = float(re['avgPx'])
                    ws['C' + str(n + result0_len)] = float(re['accFillSz'])
                    ws['D' + str(n + result0_len)] = float(re['accFillSz']) * float(re['avgPx'])
                n = n - 1
                wb.save("coin_v5.xlsx")
        else:
            result = tradeAPI.orders_history_archive(instType='SPOT', instId=coin_name + '-USDT',
                                                     before=wb[coin_name]['H1'].value)
            if len(result['data']) ==0:
                continue
            ws['H1'] = str(result['data'][0]['ordId'])
            wb.save("coin_v5.xlsx")
            n = ws.max_row
            result0_len = len(result['data'])
            for re in result['data']:
                print(coin_name + "   Updating")
                if re['side'] == "buy":
                    print("==================================")
                    print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                    ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                    ws['B' + str(n + result0_len)] = float(re['avgPx'])
                    ws['C' + str(n + result0_len)] = float(re['accFillSz']) * -1
                    ws['D' + str(n + result0_len)] = float(re['accFillSz']) * -1 * float(re['avgPx'])
                else:
                    print("==================================")
                    print(json.dumps(re, sort_keys=True, indent=4, separators=(',', ': ')))
                    ws['A' + str(n + result0_len)] = str(timestamp_datetime(int(re['fillTime'])))
                    ws['B' + str(n + result0_len)] = float(re['avgPx'])
                    ws['C' + str(n + result0_len)] = float(re['accFillSz'])
                    ws['D' + str(n + result0_len)] = float(re['accFillSz']) * float(re['avgPx'])
                n = n - 1
                wb.save("coin_v5.xlsx")

    result = tradeAPI.orders_history_archive(instType='SPOT')
    wb['BTC']['I1'] = result['data'][0]['ordId']  # 最新一条的订单信息
    wb.save("coin_v5.xlsx")

#获取单个币的买卖情况
def getcoin_one_put(coin_name):
    wb = op.load_workbook("coin_v5.xlsx")
    ws = wb[coin_name]
    colD = ws['D']
    coin_buy=0.0
    coin_sell=0.0

    for col in colD:
        if col.value is None:
            continue
        if col.value > 0.0:
            coin_buy = coin_buy + col.value
        else:
            coin_sell = coin_sell + col.value
    coin_put = coin_buy + coin_sell
    parameters = {"coin_buy": coin_buy,
                  "coin_sell": coin_sell,
                  "coin_put": coin_put
                  }
    return parameters

#更新账户的币存量 推算利润
def update_profit_all():
    parameters = init_basics()
    accountAPI = parameters["accountAPI"]
    marketAPI=parameters["marketAPI"]
    #market = marketAPI.get_ticker('BTC-USDT')
    wb = op.load_workbook("coin_v5.xlsx")
    result = accountAPI.get_account()
    for coinname in result['data'][0]['details']:
        ccy=coinname['ccy']
        if ccy not in wb.sheetnames:
            wb.create_sheet(ccy)
            wb.save("coin_v5.xlsx")
        wb = op.load_workbook("coin_v5.xlsx")
        ws=wb[ccy]
        market = marketAPI.get_ticker(ccy+'-USDT')
        if len(market['data'])==0:
            continue
        #现价 market['data'][0]['last']  存量 coinname['cashBal']
        if float(coinname['cashBal']) * float(market['data'][0]['last'])*usdt < 1.0:
            continue
        print(ccy+"更新account======================")

        parameters=getcoin_one_put(ccy)
        coin_put=parameters["coin_put"]
        # E1存放现有的存量
        ws['E1'] = float(coinname['cashBal'])
        #F1存放 目前该币的价值 USDT
        ws['F1'] = float(coinname['cashBal']) * float(market['data'][0]['last'])  # usdt价格
        #F1存放 目前该币的价值 RMB
        ws['G1'] = float(coinname['cashBal']) * float(market['data'][0]['last'])*usdt
        # E2 存放目前总营收 USDT
        ws['E2'] = float(coin_put)
        #F2存放总收益 USDT
        ws['F2']=float(coin_put)+float(coinname['cashBal']) * float(market['data'][0]['last'])
        #G2存放总收益 RMB
        ws['G2'] = float(coin_put)*usdt + float(coinname['cashBal']) * float(market['data'][0]['last'])*usdt
        wb.save("coin_v5.xlsx")

def update_profit_one(coin_name):
    parameters = init_basics()
    accountAPI = parameters["accountAPI"]
    marketAPI = parameters["marketAPI"]
    wb = op.load_workbook("coin_v5.xlsx")
    ws=wb[coin_name]
    result = accountAPI.get_account(coin_name)
    market = marketAPI.get_ticker(coin_name + '-USDT')
    print(coin_name+"  "+str(timestamp_datetime(int(market['data'][0]['ts'])))+"单价"+market['data'][0]['last'])
    if len(market['data']) == 0:
        print("数据为空，请检查！")
        return
    details=result['data'][0]['details'][0]
    parameters = getcoin_one_put(coin_name)
    coin_put = parameters["coin_put"]
    ws['E1'] = float(details['cashBal'])
    ws['F1'] = float(details['cashBal']) * float(market['data'][0]['last'])
    ws['G1'] = float(details['cashBal']) * float(market['data'][0]['last']) * usdt
    ws['E2'] = float(coin_put)
    ws['F2'] = float(coin_put) + float(details['cashBal']) * float(market['data'][0]['last'])
    ws['G2'] = float(coin_put) * usdt + float(details['cashBal']) * float(market['data'][0]['last']) * usdt
    wb.save("coin_v5.xlsx")
    # 现价 market['data'][0]['last']  存量 coinname['cashBal']
    #if float(coinname['cashBal']) * float(market['data'][0]['last']) * usdt < 1.0:
       #return

def coin_excel_profit():
    wb = op.load_workbook("coin_v5.xlsx")
    for coin_name in wb.sheetnames:
        ws = wb[coin_name]
        colD = ws['D']
        coin_buy = 0.0
        coin_sell = 0.0
        for col in colD:
            if col.value is None:
                continue
            if col.value > 0.0:
                coin_buy = coin_buy + col.value
            else:
                coin_sell = coin_sell + col.value
        coin_put = coin_buy + coin_sell
        ws['E2'] = float(coin_put)
        #E3存放已有订单的总营收RMB
        ws['E3'] = float(coin_put)*usdt
        wb.save("coin_v5.xlsx")

#查看币 现状
def check_coin_order(coin_name):
    wb = op.load_workbook("coin_v5.xlsx")
    ws=wb[coin_name]
    coin_order=PrettyTable()
    coin_order.field_names=["交易时间","单价  ","数量  ","总价  "]
    for i in range(ws.max_row):
      if  ws['A'+str(i+1)].value==None:
            continue
      coin_order.add_row([ws['A'+str(i+1)].value,ws['B'+str(i+1)].value,ws['C'+str(i+1)].value,ws['D'+str(i+1)].value])
    coin_order1 = PrettyTable()
    coin_order2 = PrettyTable()
    coin_order1.field_names = ["现有存量", "存量价值(USDT)  ", "存量价值(RMB)  "]
    coin_order1.add_row([ws['E1'].value,ws['F1'].value,ws['G1'].value])
    coin_order2.field_names = ["目前订单总营收", "总收益(USDT)  ", "总收益(RMB)  "]
    coin_order2.add_row([ws['E2'].value, ws['F2'].value, ws['G2'].value])
    coin_order2.add_row([ws['E3'].value, '', ''])
    print(coin_order)
    print(coin_order1)
    print(coin_order2)

def update():
    #1.更新订单
    update_account_order_all()
    #update_account_order_one(coin_name=)
    # 2.更新收益
    update_profit_all()
    # 3.更新excel
    coin_excel_profit()

def check_update_coin(coin_name):
    update_profit_one(coin_name)
    check_coin_order(coin_name)

def check_down():
    wb = op.load_workbook("coin_v5.xlsx")
    for coin_name in wb.sheetnames:
        ws=wb[coin_name]
        if ws['F2'].value is None:
            continue
        if float(ws['F2'].value) < 0:
            tt=float(ws['F2'].value)/float(ws['E2'].value)
            print(coin_name + "*****" + str(ws['F2'].value)+"######"+str(tt*100)+"%")
            check_update_coin(coin_name)
        #else:
            #print(coin_name + "=====" + str(ws['E2'].value))
def check_up():
    wb = op.load_workbook("coin_v5.xlsx")
    for coin_name in wb.sheetnames:
        ws=wb[coin_name]
        if ws['E2'].value is None:
            continue
        if float(ws['E2'].value) > 0:
            #tt=float(ws['E2'].value)/float(ws['E2'].value)
            print(coin_name + "*****" + str(ws['E2'].value))
            check_update_coin(coin_name)

def check_one_coin():
    coin_name = input()
    coin_name=coin_name.upper()
    wb = op.load_workbook("coin_v5.xlsx")
    if coin_name not in wb.sheetnames:
        print("查无此币记录！！！")
        return
    else:
      coin_excel_profit()
      check_update_coin(coin_name)



#balChg 账户层面的余额变动数量  billId 账单ID ts 账单创建时间t
if __name__ == '__main__':
    #init_basics()
    update()
    #update()
    #check_down()
    #check_up()
    #check_one_coin()


